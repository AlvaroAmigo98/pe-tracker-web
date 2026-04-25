from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Max, Q
from django.core.cache import cache
from .models import Company, Person, PersonSnapshot, ChangeEvent, ScrapeRun, ScrapeRunFirm, AuditLog

import json
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

audit_logger = logging.getLogger('tracker.audit')

_RATE_LIMIT     = 5      # max failed attempts
_RATE_WINDOW    = 3600   # lockout window in seconds (1 hour)


def _client_ip(request):
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    return forwarded.split(',')[0].strip() if forwarded else request.META.get('REMOTE_ADDR', '?')


class RateLimitedLoginView(LoginView):
    """LoginView that locks an IP after 5 consecutive failed attempts for 1 hour."""

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'POST':
            ip = _client_ip(request)
            attempts = cache.get(f'login_fail_{ip}', 0)
            if attempts >= _RATE_LIMIT:
                return render(request, 'tracker/login.html', {
                    'form': self.get_form_class()(),
                    'locked': True,
                })
        return super().dispatch(request, *args, **kwargs)

    def form_invalid(self, form):
        ip = _client_ip(self.request)
        key = f'login_fail_{ip}'
        attempts = cache.get(key, 0) + 1
        cache.set(key, attempts, timeout=_RATE_WINDOW)
        username = form.data.get('username', '?')
        audit_logger.warning('LOGIN_FAILED username=%s ip=%s attempts=%d', username, ip, attempts)
        try:
            AuditLog.objects.create(event_type='LOGIN_FAILED', username=username, ip_address=ip)
        except Exception:
            pass
        return super().form_invalid(form)

    def form_valid(self, form):
        ip = _client_ip(self.request)
        cache.delete(f'login_fail_{ip}')
        audit_logger.info('LOGIN user=%s ip=%s', form.get_user().username, ip)
        return super().form_valid(form)


SENIORITY_GROUP = {
    "Partner / MD": "Senior",
    "Director":     "Senior",
    "VP":           "Senior",
    "Principal":    "Junior",
    "Associate":    "Junior",
    "Analyst":      "Junior",
    "Other":        "Junior",
    "Unknown":      "Junior",
}

LOCATION_REGIONS = {
    "new york": "North America", "san francisco": "North America",
    "chicago": "North America", "boston": "North America",
    "los angeles": "North America", "houston": "North America",
    "toronto": "North America", "miami": "North America",
    "washington": "North America", "menlo park": "North America",
    "nashville": "North America", "atlanta": "North America",
    "dallas": "North America", "montreal": "North America",
    "united states": "North America", "usa": "North America",
    "canada": "North America",
    "london": "EMEA", "paris": "EMEA", "frankfurt": "EMEA",
    "amsterdam": "EMEA", "madrid": "EMEA", "milan": "EMEA",
    "stockholm": "EMEA", "zurich": "EMEA", "munich": "EMEA",
    "dubai": "EMEA", "abu dhabi": "EMEA", "luxembourg": "EMEA",
    "berlin": "EMEA", "copenhagen": "EMEA", "oslo": "EMEA",
    "helsinki": "EMEA", "warsaw": "EMEA", "vienna": "EMEA",
    "brussels": "EMEA", "dublin": "EMEA", "lisbon": "EMEA",
    "uk": "EMEA", "united kingdom": "EMEA", "germany": "EMEA",
    "france": "EMEA", "sweden": "EMEA", "netherlands": "EMEA",
    "spain": "EMEA", "italy": "EMEA", "switzerland": "EMEA",
    "denmark": "EMEA", "norway": "EMEA", "finland": "EMEA",
    "poland": "EMEA", "austria": "EMEA", "belgium": "EMEA",
    "ireland": "EMEA", "portugal": "EMEA", "middle east": "EMEA",
    "hong kong": "APAC", "singapore": "APAC", "tokyo": "APAC",
    "shanghai": "APAC", "beijing": "APAC", "sydney": "APAC",
    "melbourne": "APAC", "seoul": "APAC", "mumbai": "APAC",
    "bangalore": "APAC", "delhi": "APAC", "taipei": "APAC",
    "china": "APAC", "japan": "APAC", "australia": "APAC",
    "india": "APAC", "south korea": "APAC", "taiwan": "APAC",
}


def infer_region(location: str) -> str:
    if not location or location == "N/A":
        return "Unknown"
    loc = location.lower()
    for keyword, region in LOCATION_REGIONS.items():
        if keyword in loc:
            return region
    return "Unknown"


def infer_seniority_group(seniority: str) -> str:
    return SENIORITY_GROUP.get(seniority, "Junior")


def infer_function_web(title: str) -> str:
    if not title or title == "N/A":
        return "Unknown"
    t = title.lower()

    if any(x in t for x in [
        "human resources", " hr,", " hr ", "talent acquisition",
        "recruiting", "recruitment", "office manager", "facilities",
        "procurement", "events ", "marketing", "communications",
        "public relations", " pr ", "legal counsel", "general counsel",
        "compliance officer", "risk officer", "audit", "tax ",
        "accounting", "treasury", "fund admin", "fund finance",
        "fund operations", "portfolio reporting", "valuations",
        "luxembourg operations", "it support", "cyber",
        "information security", "esg", "sustainability",
        "investor relations", "capital formation",
    ]):
        return "Operations"

    if any(x in t for x in [
        "senior adviser", "senior advisor", "adviser", "advisor",
        "board member", "board director", "chairman", "chairwoman",
        "co-founder", "founding partner", "executive chairman",
        "executive in residence", "entrepreneur in residence",
        "venture partner", "operating partner", "executive advisor",
    ]):
        return "Advisory"

    if any(x in t for x in [
        "buyout", "private equity", " pe ", "leveraged",
        "lbo", "growth equity", "growth capital",
    ]):
        return "Buyout / PE"

    if any(x in t for x in [
        "infrastructure", "infra", "transport", "energy transition",
        "renewable", "utilities", "power",
    ]):
        return "Infrastructure"

    if any(x in t for x in [
        "real estate", "property", "realty", "reit", "real assets",
    ]):
        return "Real Estate"

    if any(x in t for x in [
        "credit", "debt", "lending", "fixed income",
        "distressed", "mezzanine", "direct lending",
        "structured finance", "clo",
    ]):
        return "Credit / Debt"

    if any(x in t for x in [
        "venture", " vc ", "early stage", "seed",
    ]):
        return "Venture Capital"

    if any(x in t for x in [
        "managing director", "director", "partner", "principal",
        "associate", "analyst", "vice president", " vp",
        "head of", "co-head", "investment", "portfolio",
        "deal", "transaction", "m&a", "origination",
        "execution", "coverage", "sector",
    ]):
        return "Investment (General)"

    return "Other"


def _get_latest_snapshots_qs(company=None):
    """Return latest snapshot per person using DISTINCT ON (PostgreSQL).
    Reduces DB load from loading all historical rows to only the latest per person.
    """
    qs = PersonSnapshot.objects.order_by('person_id', '-scraped_at').distinct('person_id')
    if company:
        qs = PersonSnapshot.objects.filter(
            person__company=company
        ).order_by('person_id', '-scraped_at').distinct('person_id')
    return qs


def _make_excel_response(filename, headers, rows, col_widths):
    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(col, 18)
    for row_num, row_data in enumerate(rows, start=2):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)
    ws.freeze_panes = "A2"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def dashboard(request):
    today            = date.today()
    latest_run       = ScrapeRun.objects.order_by('-ran_at').first()
    companies        = Company.objects.order_by('name')
    company_filter      = request.GET.getlist('company')
    region_filter       = request.GET.getlist('region')
    group_filter        = request.GET.getlist('group')
    function_filter     = request.GET.getlist('function')
    change_type_filter  = request.GET.getlist('change_type')
    export              = request.GET.get('export', '')
    days                = request.GET.get('days', '')
    date_from           = request.GET.get('date_from', '')
    date_to             = request.GET.get('date_to', '')

    events_qs = ChangeEvent.objects.select_related(
        'person', 'person__company'
    ).order_by('-detected_at', '-id')

    if days and days.isdigit():
        cutoff = today - timedelta(days=int(days))
        events_qs = events_qs.filter(detected_at__gte=cutoff)
    if date_from:
        try:
            events_qs = events_qs.filter(detected_at__gte=date.fromisoformat(date_from))
        except ValueError:
            date_from = ''
    if date_to:
        try:
            events_qs = events_qs.filter(detected_at__lte=date.fromisoformat(date_to))
        except ValueError:
            date_to = ''

    events = list(events_qs)

    person_ids = list({e.person_id for e in events})

    latest_snapshot_by_person = {}
    if person_ids:
        latest_snapshots_qs = PersonSnapshot.objects.filter(
            person_id__in=person_ids
        ).select_related('person').order_by('person_id', '-scraped_at', '-id')

        for snap in latest_snapshots_qs:
            if snap.person_id not in latest_snapshot_by_person:
                latest_snapshot_by_person[snap.person_id] = snap

    seven_days_ago = today - timedelta(days=7)
    for e in events:
        latest_snap = latest_snapshot_by_person.get(e.person_id)
        e.region = infer_region(latest_snap.location if latest_snap else '')
        e.function = infer_function_web(e.new_title or e.previous_title or '')
        e.seniority_group = infer_seniority_group(
            latest_snap.seniority if latest_snap else ''
        )
        dt = e.detected_at
        if hasattr(dt, 'date'):
            dt = dt.date()
        e.is_new = dt >= seven_days_ago

    if company_filter:
        events = [e for e in events if e.person.company.name in company_filter]
    if region_filter:
        events = [e for e in events if e.region in region_filter]
    if group_filter:
        events = [e for e in events if e.seniority_group in group_filter]
    if function_filter:
        events = [e for e in events if e.function in function_filter]
    if change_type_filter:
        def _matches_change_type(e):
            for ct in change_type_filter:
                if ct == 'hire' and e.event_type == 'hire':
                    return True
                if ct == 'leaver' and e.event_type == 'leaver':
                    return True
                if ct == 'role_change' and e.event_type in ('promotion', 'role_change'):
                    return True
            return False
        events = [e for e in events if _matches_change_type(e)]

    hires      = [e for e in events if e.event_type == 'hire']
    leavers    = [e for e in events if e.event_type == 'leaver']
    promotions = [e for e in events if e.event_type in ('promotion', 'role_change')]

    if export == 'excel':
        audit_logger.info('EXPORT view=dashboard user=%s ip=%s rows=%d',
                          request.user.username, _client_ip(request), len(events))
        headers = [
            'Name', 'Company', 'Function', 'Change Type', 'Previous Title',
            'New Title', 'Previous Level', 'New Level', 'Region',
            'Seniority Group', 'Detected At'
        ]
        rows = [
            [
                e.person.full_name, e.person.company.name, e.function,
                e.event_type, e.previous_title or '—', e.new_title or '—',
                e.previous_level or '—', e.new_level or '—',
                e.region, e.seniority_group, str(e.detected_at),
            ]
            for e in events
        ]
        widths = {1:28,2:24,3:22,4:18,5:32,6:32,7:16,8:16,9:16,10:18,11:16}
        return _make_excel_response('historical_changes.xlsx', headers, rows, widths)

    # Watchlist activity since last visit (Feature 10)
    watchlist = list(request.session.get('watchlist', []))
    last_visit_str = request.session.get('last_visit')
    watchlist_activity = []
    if watchlist and last_visit_str:
        try:
            last_visit_dt = date.fromisoformat(last_visit_str)
            watchlist_activity = list(
                ChangeEvent.objects.filter(
                    person__company_id__in=watchlist,
                    detected_at__gt=last_visit_dt,
                ).select_related('person', 'person__company').order_by('-detected_at')[:10]
            )
        except ValueError:
            pass
    request.session['last_visit'] = str(today)

    return render(request, 'tracker/dashboard.html', {
        'latest_run':        latest_run,
        'companies':         companies,
        'company_filter':    company_filter,
        'region_filter':     region_filter,
        'group_filter':      group_filter,
        'function_filter':     function_filter,
        'change_type_filter':  change_type_filter,
        'date_from':           date_from,
        'date_to':             date_to,
        'events':              events,
        'hires':             hires,
        'leavers':           leavers,
        'promotions':        promotions,
        'days':              days,
        'watchlist':         watchlist,
        'watchlist_activity': watchlist_activity,
    })


@login_required
def people(request):
    search           = request.GET.get('q', '')
    company_filter   = request.GET.getlist('company')
    seniority_filter = request.GET.getlist('seniority')
    group_filter     = request.GET.getlist('group')
    region_filter    = request.GET.getlist('region')
    function_filter  = request.GET.getlist('function')
    last_seen_filter = request.GET.get('last_seen', '')
    export           = request.GET.get('export', '')

    companies   = Company.objects.order_by('name')
    seniorities = PersonSnapshot.objects.values_list(
        'seniority', flat=True
    ).distinct().order_by('seniority')

    # DISTINCT ON (person_id) — returns only the latest snapshot per person.
    # Replaces the old full-table scan + Python dedup loop.
    people_qs = PersonSnapshot.objects.order_by(
        'person_id', '-scraped_at'
    ).distinct('person_id').select_related('person', 'person__company')

    people = []
    for snap in people_qs:
        snap.region = infer_region(snap.location or '')
        snap.seniority_group = infer_seniority_group(snap.seniority or '')
        snap.function = infer_function_web(snap.job_title or '')
        people.append(snap)

    people.sort(key=lambda p: (p.person.company.name, p.person.full_name))

    if search:
        people = [
            p for p in people if
            search.lower() in p.person.full_name.lower() or
            search.lower() in (p.job_title or '').lower()
        ]
    if company_filter:
        people = [p for p in people if p.person.company.name in company_filter]
    if seniority_filter:
        people = [p for p in people if p.seniority in seniority_filter]
    if group_filter:
        people = [p for p in people if p.seniority_group in group_filter]
    if region_filter:
        people = [p for p in people if p.region in region_filter]
    if function_filter:
        people = [p for p in people if p.function in function_filter]
    if last_seen_filter:
        cutoff = date.today() - timedelta(days=int(last_seen_filter))
        people = [p for p in people if p.scraped_at >= cutoff]

    if export == 'excel':
        audit_logger.info('EXPORT view=people user=%s ip=%s rows=%d',
                          request.user.username, _client_ip(request), len(people))
        headers = [
            'Name', 'Company', 'Title', 'Seniority', 'Group',
            'Function', 'Region', 'Location', 'Last Seen'
        ]
        rows = [
            [
                snap.person.full_name, snap.person.company.name,
                snap.job_title or '—', snap.seniority or '—',
                snap.seniority_group, snap.function, snap.region,
                snap.location or '—', str(snap.scraped_at),
            ]
            for snap in people
        ]
        widths = {i: 20 for i in range(1, 10)}
        return _make_excel_response('pe_tracker_export.xlsx', headers, rows, widths)

    return render(request, 'tracker/people.html', {
        'people':           people,
        'companies':        companies,
        'seniorities':      seniorities,
        'search':           search,
        'company_filter':   company_filter,
        'seniority_filter': seniority_filter,
        'group_filter':     group_filter,
        'region_filter':    region_filter,
        'function_filter':  function_filter,
        'last_seen_filter': last_seen_filter,
    })


@login_required
def firms(request):
    sort_by = request.GET.get('sort', 'leavers')
    valid_sorts = {'leavers', 'hires', 'promotions', 'headcount', 'name', 'activity'}
    if sort_by not in valid_sorts:
        sort_by = 'leavers'

    latest_date = PersonSnapshot.objects.aggregate(d=Max('scraped_at'))['d']

    headcount_map = {}
    if latest_date:
        headcount_qs = PersonSnapshot.objects.filter(
            scraped_at=latest_date
        ).values('person__company_id').annotate(headcount=Count('id'))
        headcount_map = {row['person__company_id']: row['headcount'] for row in headcount_qs}

    event_qs = ChangeEvent.objects.values('person__company_id').annotate(
        hires=Count('id', filter=Q(event_type='hire')),
        leavers=Count('id', filter=Q(event_type='leaver')),
        promotions=Count('id', filter=Q(event_type__in=['promotion', 'role_change'])),
    )
    event_map = {row['person__company_id']: row for row in event_qs}

    companies = Company.objects.order_by('name')
    firm_stats = []
    for company in companies:
        stats = event_map.get(company.id, {})
        hires      = stats.get('hires', 0)
        leavers    = stats.get('leavers', 0)
        promotions = stats.get('promotions', 0)
        firm_stats.append({
            'company':    company,
            'headcount':  headcount_map.get(company.id, 0),
            'hires':      hires,
            'leavers':    leavers,
            'promotions': promotions,
            'activity':   hires + leavers + promotions,
        })

    if sort_by == 'name':
        firm_stats.sort(key=lambda x: x['company'].name)
    else:
        firm_stats.sort(key=lambda x: x[sort_by], reverse=True)

    totals = {
        'headcount':  sum(f['headcount']  for f in firm_stats),
        'hires':      sum(f['hires']      for f in firm_stats),
        'leavers':    sum(f['leavers']    for f in firm_stats),
        'promotions': sum(f['promotions'] for f in firm_stats),
    }

    watchlist = list(request.session.get('watchlist', []))
    return render(request, 'tracker/firms.html', {
        'firm_stats':  firm_stats,
        'sort_by':     sort_by,
        'totals':      totals,
        'latest_date': latest_date,
        'watchlist':   watchlist,
    })


@login_required
def firm_detail(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    tab     = request.GET.get('tab', 'team')
    search  = request.GET.get('q', '')
    export  = request.GET.get('export', '')

    # Latest snapshot per person for this firm
    team_qs = PersonSnapshot.objects.filter(
        person__company=company
    ).order_by('person_id', '-scraped_at').distinct('person_id').select_related('person')

    snap_by_person = {}
    team = []
    for snap in team_qs:
        snap.region = infer_region(snap.location or '')
        snap.seniority_group = infer_seniority_group(snap.seniority or '')
        snap.function = infer_function_web(snap.job_title or '')
        snap_by_person[snap.person_id] = snap
        team.append(snap)

    team.sort(key=lambda p: p.person.full_name)

    # All change events for this firm
    events_qs = ChangeEvent.objects.filter(
        person__company=company
    ).select_related('person').order_by('-detected_at', '-id')

    events = list(events_qs)
    for e in events:
        snap = snap_by_person.get(e.person_id)
        e.region = infer_region(snap.location if snap else '')
        e.function = infer_function_web(e.new_title or e.previous_title or '')
        e.seniority_group = infer_seniority_group(snap.seniority if snap else '')

    hires      = [e for e in events if e.event_type == 'hire']
    leavers    = [e for e in events if e.event_type == 'leaver']
    promotions = [e for e in events if e.event_type in ('promotion', 'role_change')]

    # Apply search to the active tab
    if search:
        def name_match(name, t1='', t2=''):
            q = search.lower()
            return q in name.lower() or q in t1.lower() or q in t2.lower()

        team = [p for p in team if
            name_match(p.person.full_name, p.job_title or '')]
        if tab == 'hires':
            hires = [e for e in hires if
                name_match(e.person.full_name, e.new_title or '', e.previous_title or '')]
        elif tab == 'leavers':
            leavers = [e for e in leavers if
                name_match(e.person.full_name, e.new_title or '', e.previous_title or '')]
        elif tab == 'promotions':
            promotions = [e for e in promotions if
                name_match(e.person.full_name, e.new_title or '', e.previous_title or '')]

    # Excel exports
    if export == 'team':
        headers = ['Name', 'Title', 'Seniority', 'Group', 'Function', 'Region', 'Location', 'Last Seen']
        rows = [
            [p.person.full_name, p.job_title or '—', p.seniority or '—',
             p.seniority_group, p.function, p.region, p.location or '—', str(p.scraped_at)]
            for p in team
        ]
        return _make_excel_response(f'{company.name}_team.xlsx', headers, rows, {i: 22 for i in range(1, 9)})

    if export == 'events':
        headers = ['Name', 'Change Type', 'Previous Title', 'New Title', 'Previous Level', 'New Level', 'Function', 'Region', 'Date']
        source = {'hires': hires, 'leavers': leavers, 'promotions': promotions}.get(tab, events)
        rows = [
            [e.person.full_name, e.event_type, e.previous_title or '—', e.new_title or '—',
             e.previous_level or '—', e.new_level or '—', e.function, e.region, str(e.detected_at)]
            for e in source
        ]
        return _make_excel_response(f'{company.name}_{tab}.xlsx', headers, rows, {i: 22 for i in range(1, 10)})

    watchlist = list(request.session.get('watchlist', []))
    return render(request, 'tracker/firm_detail.html', {
        'company':    company,
        'tab':        tab,
        'search':     search,
        'team':       team,
        'hires':      hires,
        'leavers':    leavers,
        'promotions': promotions,
        'all_events': events,
        'is_watched': company.id in watchlist,
    })


@login_required
def signals(request):
    today = date.today()
    senior_levels = ['Partner / MD', 'Director', 'VP']

    # ── Feature 1: Cascade departure alerts ───────────────────────────────────
    # Senior (Partner/MD or Director) leaves → 2+ more people leave same firm within 8 weeks
    all_leaver_list = list(ChangeEvent.objects.filter(
        event_type='leaver',
    ).select_related('person', 'person__company').order_by('detected_at'))

    cascade_alerts = []
    seen_trigger_ids = set()
    for trigger in all_leaver_list:
        if trigger.previous_level not in ['Partner / MD', 'Director']:
            continue
        if trigger.id in seen_trigger_ids:
            continue
        end_window = trigger.detected_at + timedelta(days=56)
        followers = [
            e for e in all_leaver_list
            if e.person.company_id == trigger.person.company_id
            and e.person_id != trigger.person_id
            and trigger.detected_at <= e.detected_at <= end_window
        ]
        if len(followers) >= 2:
            seen_trigger_ids.add(trigger.id)
            cascade_alerts.append({
                'firm': trigger.person.company.name,
                'company_id': trigger.person.company_id,
                'trigger_name': trigger.person.full_name,
                'trigger_title': trigger.previous_title or trigger.previous_level or '—',
                'trigger_date': trigger.detected_at,
                'follower_names': ', '.join(f.person.full_name for f in followers[:4])
                                  + (f' +{len(followers)-4} more' if len(followers) > 4 else ''),
                'count': len(followers),
            })

    cascade_alerts.sort(key=lambda x: -x['count'])
    seen_cascade_firms = set()
    deduped_cascades = []
    for a in cascade_alerts:
        if a['firm'] not in seen_cascade_firms:
            seen_cascade_firms.add(a['firm'])
            deduped_cascades.append(a)
    cascade_alerts = deduped_cascades

    # ── Feature 2: Lift-out / Spinout detector ────────────────────────────────
    # 4+ people leave same firm within 21 days
    liftout_signals = []
    seen_windows = set()
    for anchor in all_leaver_list:
        end_window = anchor.detected_at + timedelta(days=21)
        cluster = [
            e for e in all_leaver_list
            if e.person.company_id == anchor.person.company_id
            and anchor.detected_at <= e.detected_at <= end_window
        ]
        if len(cluster) >= 4:
            key = (anchor.person.company_id, anchor.detected_at)
            if key not in seen_windows:
                seen_windows.add(key)
                liftout_signals.append({
                    'firm': anchor.person.company.name,
                    'company_id': anchor.person.company_id,
                    'count': len(cluster),
                    'start': cluster[0].detected_at,
                    'end': cluster[-1].detected_at,
                    'names': ', '.join(e.person.full_name for e in cluster[:5])
                             + (f' +{len(cluster)-5} more' if len(cluster) > 5 else ''),
                })

    # deduplicate overlapping windows per firm, keep highest count
    liftout_signals.sort(key=lambda x: (-x['count'], x['start']))
    seen_liftout_firms = set()
    deduped_liftouts = []
    for s in liftout_signals:
        if s['firm'] not in seen_liftout_firms:
            seen_liftout_firms.add(s['firm'])
            deduped_liftouts.append(s)
    liftout_signals = deduped_liftouts[:10]

    # ── Feature 4: Firm health score (rolling 90-day) ─────────────────────────
    ninety_days_ago = today - timedelta(days=90)
    latest_date = PersonSnapshot.objects.aggregate(d=Max('scraped_at'))['d']
    date_90 = PersonSnapshot.objects.filter(
        scraped_at__lte=ninety_days_ago
    ).aggregate(d=Max('scraped_at'))['d']

    current_hc = {}
    if latest_date:
        for row in PersonSnapshot.objects.filter(
            scraped_at=latest_date
        ).values('person__company__name').annotate(n=Count('id')):
            current_hc[row['person__company__name']] = row['n']

    past_hc = {}
    if date_90:
        for row in PersonSnapshot.objects.filter(
            scraped_at=date_90
        ).values('person__company__name').annotate(n=Count('id')):
            past_hc[row['person__company__name']] = row['n']

    sleavers_90 = {}
    for row in ChangeEvent.objects.filter(
        event_type='leaver',
        detected_at__gte=ninety_days_ago,
        previous_level__in=senior_levels,
    ).values('person__company__name').annotate(n=Count('id')):
        sleavers_90[row['person__company__name']] = row['n']

    health_scores = []
    for firm_name, curr in current_hc.items():
        prev = past_hc.get(firm_name, curr)
        net = curr - prev
        sl = sleavers_90.get(firm_name, 0)
        avg_hc = (curr + prev) / 2 or 1
        attrition_rate = round(sl / avg_hc * 100, 1)
        score = net - sl * 2
        health_scores.append({
            'firm': firm_name,
            'current': curr,
            'net': net,
            'senior_leavers': sl,
            'attrition_rate': attrition_rate,
            'score': score,
        })

    health_scores.sort(key=lambda x: x['score'])

    worst_20 = health_scores[:20]
    health_chart_json = json.dumps({
        'labels': [h['firm'] for h in worst_20],
        'net': [h['net'] for h in worst_20],
        'colors': [
            'rgba(231,76,60,0.85)' if h['net'] < 0 else 'rgba(46,204,113,0.85)'
            for h in worst_20
        ],
    })

    # ── Feature 5: Seniority pyramid (top 15 firms by headcount) ──────────────
    SENIORITY_LEVELS = ['Partner / MD', 'Director', 'VP', 'Principal', 'Associate', 'Analyst']
    PYRAMID_COLORS   = ['#1a1a2e', '#2d4a8a', '#4a7fc1', '#7aaddb', '#a8cce8', '#d0e8f5']

    top_firm_names = [f for f, _ in sorted(current_hc.items(), key=lambda x: -x[1])[:15]]

    pyramid_raw = {}
    if latest_date and top_firm_names:
        for row in PersonSnapshot.objects.filter(
            scraped_at=latest_date,
            person__company__name__in=top_firm_names,
        ).values('person__company__name', 'seniority').annotate(n=Count('id')):
            firm = row['person__company__name']
            pyramid_raw.setdefault(firm, {})[row['seniority']] = row['n']

    pyramid_chart_json = json.dumps({
        'labels': top_firm_names,
        'datasets': [
            {
                'label': level,
                'data': [pyramid_raw.get(f, {}).get(level, 0) for f in top_firm_names],
                'backgroundColor': PYRAMID_COLORS[i],
            }
            for i, level in enumerate(SENIORITY_LEVELS)
        ],
    })

    return render(request, 'tracker/signals.html', {
        'cascade_alerts':     cascade_alerts[:10],
        'liftout_signals':    liftout_signals,
        'health_scores':      health_scores,
        'health_chart_json':  health_chart_json,
        'pyramid_chart_json': pyramid_chart_json,
        'today':              today,
    })


@login_required
def search(request):
    q = request.GET.get('q', '').strip()
    firm_results = []
    person_results = []
    event_results = []
    if q:
        firm_results = list(Company.objects.filter(name__icontains=q).order_by('name')[:20])
        person_results = list(
            PersonSnapshot.objects.filter(
                Q(person__full_name__icontains=q) | Q(job_title__icontains=q)
            ).order_by('person_id', '-scraped_at').distinct('person_id')
            .select_related('person', 'person__company')[:30]
        )
        for snap in person_results:
            snap.region = infer_region(snap.location or '')
            snap.seniority_group = infer_seniority_group(snap.seniority or '')
            snap.function = infer_function_web(snap.job_title or '')
        event_results = list(
            ChangeEvent.objects.filter(
                Q(person__full_name__icontains=q) |
                Q(new_title__icontains=q) |
                Q(previous_title__icontains=q)
            ).select_related('person', 'person__company').order_by('-detected_at')[:20]
        )
    return render(request, 'tracker/search_results.html', {
        'q':              q,
        'firm_results':   firm_results,
        'person_results': person_results,
        'event_results':  event_results,
    })


@login_required
def watchlist_toggle(request, company_id):
    if request.method != 'POST':
        return redirect('firms')
    watchlist = list(request.session.get('watchlist', []))
    if company_id in watchlist:
        watchlist.remove(company_id)
    else:
        watchlist.append(company_id)
    request.session['watchlist'] = watchlist
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or '/'
    return redirect(next_url)


@login_required
def api_search(request):
    q = request.GET.get('q', '').strip()
    results = {'firms': [], 'people': []}
    if len(q) >= 2:
        firms = Company.objects.filter(name__icontains=q).order_by('name')[:6]
        results['firms'] = [{'id': f.id, 'name': f.name} for f in firms]
        people_qs = PersonSnapshot.objects.filter(
            Q(person__full_name__icontains=q) | Q(job_title__icontains=q)
        ).order_by('person_id', '-scraped_at').distinct('person_id').select_related('person', 'person__company')[:6]
        results['people'] = [
            {'name': p.person.full_name, 'firm': p.person.company.name, 'title': p.job_title or ''}
            for p in people_qs
        ]
    return JsonResponse(results)


@login_required
def firm_report(request, company_id):
    company = get_object_or_404(Company, id=company_id)

    team_qs = PersonSnapshot.objects.filter(
        person__company=company
    ).order_by('person_id', '-scraped_at').distinct('person_id').select_related('person')

    snap_by_person = {}
    team = []
    for snap in team_qs:
        snap.region = infer_region(snap.location or '')
        snap.seniority_group = infer_seniority_group(snap.seniority or '')
        snap.function = infer_function_web(snap.job_title or '')
        snap_by_person[snap.person_id] = snap
        team.append(snap)
    team.sort(key=lambda p: p.person.full_name)

    events_qs = ChangeEvent.objects.filter(
        person__company=company
    ).select_related('person').order_by('-detected_at', '-id')
    events = list(events_qs)
    for e in events:
        snap = snap_by_person.get(e.person_id)
        e.region = infer_region(snap.location if snap else '')
        e.function = infer_function_web(e.new_title or e.previous_title or '')
        e.seniority_group = infer_seniority_group(snap.seniority if snap else '')

    hires      = [e for e in events if e.event_type == 'hire']
    leavers    = [e for e in events if e.event_type == 'leaver']
    promotions = [e for e in events if e.event_type in ('promotion', 'role_change')]
    senior_leavers = [e for e in leavers if e.seniority_group == 'Senior']

    return render(request, 'tracker/firm_report.html', {
        'company':        company,
        'team':           team,
        'hires':          hires,
        'leavers':        leavers,
        'promotions':     promotions,
        'senior_leavers': senior_leavers,
        'generated_at':   date.today(),
    })


@_superuser_required
def scrape_logs(request):
    runs = list(ScrapeRun.objects.order_by('-ran_at')[:52])
    run_ids = [r.id for r in runs]
    firms = ScrapeRunFirm.objects.filter(run_id__in=run_ids).order_by('firm_name')
    firms_by_run = {}
    for f in firms:
        firms_by_run.setdefault(f.run_id, []).append(f)
    for r in runs:
        r.firm_details = firms_by_run.get(r.id, [])
    return render(request, 'tracker/scrape_logs.html', {'runs': runs})


# ─────────────────────────────────────────────────────────────────────────────
# Profile
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def profile(request):
    user = request.user
    message = None
    error = None

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_details':
            first_name = request.POST.get('first_name', '').strip()
            last_name  = request.POST.get('last_name', '').strip()
            email      = request.POST.get('email', '').strip()
            user.first_name = first_name
            user.last_name  = last_name
            user.email      = email
            user.save()
            message = 'Profile updated.'

        elif action == 'change_password':
            current  = request.POST.get('current_password', '')
            new_pw   = request.POST.get('new_password', '')
            confirm  = request.POST.get('confirm_password', '')
            if not user.check_password(current):
                error = 'Current password is incorrect.'
            elif len(new_pw) < 8:
                error = 'New password must be at least 8 characters.'
            elif new_pw != confirm:
                error = 'New passwords do not match.'
            else:
                user.set_password(new_pw)
                user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)
                message = 'Password changed successfully.'

    recent_logins = AuditLog.objects.filter(
        username=user.username, event_type='LOGIN'
    ).order_by('-created_at')[:10]

    return render(request, 'tracker/profile.html', {
        'profile_user': user,
        'message':      message,
        'error':        error,
        'recent_logins': recent_logins,
    })


# ─────────────────────────────────────────────────────────────────────────────
# User admin (superuser only)
# ─────────────────────────────────────────────────────────────────────────────

def _superuser_required(view_fn):
    from functools import wraps
    @wraps(view_fn)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    return wrapper


@_superuser_required
def user_admin(request):
    from django.contrib.auth.models import User
    message = None
    error   = None

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_user':
            username = request.POST.get('username', '').strip()
            email    = request.POST.get('email', '').strip()
            password = request.POST.get('password', '')
            if not username or not password:
                error = 'Username and password are required.'
            elif User.objects.filter(username=username).exists():
                error = f'Username "{username}" already exists.'
            elif len(password) < 8:
                error = 'Password must be at least 8 characters.'
            else:
                User.objects.create_user(username=username, email=email, password=password)
                message = f'User "{username}" created.'

        elif action == 'toggle_user':
            uid  = request.POST.get('user_id')
            user = User.objects.filter(id=uid).exclude(id=request.user.id).first()
            if user:
                user.is_active = not user.is_active
                user.save()
                state = 'activated' if user.is_active else 'deactivated'
                message = f'User "{user.username}" {state}.'

        elif action == 'delete_user':
            uid  = request.POST.get('user_id')
            user = User.objects.filter(id=uid).exclude(id=request.user.id).first()
            if user:
                name = user.username
                user.delete()
                message = f'User "{name}" deleted.'

    users       = User.objects.order_by('-date_joined')
    audit_logs  = AuditLog.objects.order_by('-created_at')[:200]

    return render(request, 'tracker/user_admin.html', {
        'users':      users,
        'audit_logs': audit_logs,
        'message':    message,
        'error':      error,
    })
